from openpyxl.styles import Font, PatternFill

def format_sheet(ws, money_columns=None, highlight_low_stock=False):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws.freeze_panes = "A2"

    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
        ws.column_dimensions[column[0].column_letter].width = max_length + 3

    if money_columns:
        for col in money_columns:
            for cell in ws[col][1:]:
                cell.number_format = '"$"#,##0.00'

    if highlight_low_stock:
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in ws.iter_rows(min_row=2):
            if row[-1].value == "LOW STOCK":
                for cell in row:
                    cell.fill = red_fill
